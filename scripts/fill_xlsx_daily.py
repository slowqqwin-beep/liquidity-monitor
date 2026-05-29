"""
Auto-fill v5.1 cleaned_v2.xlsx daily template from series.json.
Reads FRED + Yahoo data and fills Today + 20d delta columns.
Outputs a dated copy to daily_logs/.

Usage:
    python scripts/fill_xlsx_daily.py
    python scripts/fill_xlsx_daily.py --date 2026-05-27
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
WORKSPACE_ROOT = REPO_ROOT.parent  # D:\liquidity-dashboard
DATA_DIR = REPO_ROOT / "data"
# Default: the user's working template in workspace root
# CI fallback: template_cleaned_v2.xlsx inside the repo
if (WORKSPACE_ROOT / "v5.1_看板每日填表_模板_cleaned_v2.xlsx").exists():
    TEMPLATE_XLSX = WORKSPACE_ROOT / "v5.1_看板每日填表_模板_cleaned_v2.xlsx"
else:
    TEMPLATE_XLSX = REPO_ROOT / "template_cleaned_v2.xlsx"
OUTPUT_DIR = REPO_ROOT / "daily_logs"
SERIES_PATH = DATA_DIR / "series.json"

# FRED/Yahoo series IDs
DGS2, DGS5, DGS10, DGS30 = "DGS2", "DGS5", "DGS10", "DGS30"
DFII10 = "DFII10"
T5YIFR = "T5YIFR"
T10YIE = "T10YIE"
SOFR, IORB, EFFR = "SOFR", "IORB", "EFFR"
DFEDTARL = "DFEDTARL"
WRESBAL = "WRESBAL"
BAMLH0A0HYM2 = "BAMLH0A0HYM2"
BAMLC0A0CM = "BAMLC0A0CM"
BAMLEMCBPIOAS = "BAMLEMCBPIOAS"
MORTGAGE30US = "MORTGAGE30US"
SPY, HYG, FXY = "SPY", "HYG", "FXY"


def load_data():
    if not SERIES_PATH.exists():
        print(f"[ERROR] {SERIES_PATH} not found", file=sys.stderr)
        sys.exit(1)
    with SERIES_PATH.open() as f:
        return json.load(f)


def last(s: list) -> float | None:
    return s[-1]["value"] if s else None


def ago(s: list, n: int = 20) -> float | None:
    if not s or len(s) < n + 1:
        return None
    return s[-n - 1]["value"]


def chg(s: list, n: int = 20) -> float | None:
    a, b = last(s), ago(s, n)
    return b - a if (a is not None and b is not None) else None


def bp(v: float | None) -> float | None:
    return round(v * 100, 1) if v is not None else None


def ret_pct(s: list, n: int) -> float | None:
    a = ago(s, n)
    b = last(s)
    return round((b - a) / a * 100, 1) if (a is not None and b is not None and a != 0) else None


def ma_n(s: list, n: int) -> float | None:
    if not s or len(s) < n:
        return None
    return round(sum(d["value"] for d in s[-n:]) / n, 1)


def build_spread_series(data, id1, id2):
    s1, s2 = data.get(id1, []), data.get(id2, [])
    if not s1 or not s2:
        return []
    d1 = {d["date"]: d["value"] for d in s1}
    d2 = {d["date"]: d["value"] for d in s2}
    return [{"date": dt, "value": d1[dt] - d2[dt]} for dt in sorted(set(d1) & set(d2))]


def compute_curve_regime(data) -> str:
    d2, d10 = data.get(DGS2, []), data.get(DGS10, [])
    if not d2 or not d10:
        return "N/A"
    y2, y10 = last(d2) or 0, last(d10) or 0
    spread = y10 - y2
    c2, c10 = chg(d2, 20) or 0, chg(d10, 20) or 0
    cs = c10 - c2
    direction = "Steepening" if cs > 0.03 else ("Flattening" if cs < -0.03 else "Stable")
    bias = "Bear" if c2 > 0.03 else ("Bull" if c2 < -0.03 else "")
    if spread > 1.0 and cs > 0.03 and c2 < -0.02:
        return "Steep-Steepening"
    return f"{bias} {direction}" if bias else direction


def fill_xlsx(ws, data, fill_date):
    parsed = datetime.strptime(fill_date, "%Y-%m-%d")
    ws["B4"] = parsed
    ws["B4"].number_format = "YYYY-MM-DD"
    ws["B5"] = parsed.strftime("%a").upper()

    # Pre-compute derived series
    sp_5s30s = build_spread_series(data, DGS30, DGS5)
    sp_10s30s = build_spread_series(data, DGS30, DGS10)
    sp_2s10s = build_spread_series(data, DGS10, DGS2)
    sp_2s30s = build_spread_series(data, DGS30, DGS2)
    sp_5s10s = build_spread_series(data, DGS10, DGS5)
    sp_mort = build_spread_series(data, MORTGAGE30US, DGS10)

    # Row → (today_value, delta_20d_value)
    fills = {}

    # C-Curve
    fills[11] = (bp(last(sp_5s30s)), bp(chg(sp_5s30s)))
    fills[12] = (bp(last(sp_10s30s)), bp(chg(sp_10s30s)))
    fills[13] = (bp(last(sp_2s10s)), bp(chg(sp_2s10s)))
    fills[15] = (bp(last(sp_2s30s)), bp(chg(sp_2s30s)))
    fills[16] = (bp(last(sp_5s10s)), bp(chg(sp_5s10s)))

    # C-RealYield
    fills[18] = (round(last(data.get(DFII10, [])) or 0, 2), bp(chg(data.get(DFII10, []))))

    # C-Nominal
    fills[19] = (round(last(data.get(DGS10, [])) or 0, 2), bp(chg(data.get(DGS10, []))))
    fills[20] = (round(last(data.get(DGS30, [])) or 0, 2), bp(chg(data.get(DGS30, []))))

    # C-InflationBE 10Y
    fills[29] = (round(last(data.get(T10YIE, [])) or 0, 2), bp(chg(data.get(T10YIE, []))))

    # C-InflationFwd
    fills[32] = (round(last(data.get(T5YIFR, [])) or 0, 2), bp(chg(data.get(T5YIFR, []))))

    # A-MMSpreads
    fills[38] = (
        round(((last(data.get(SOFR, [])) or 0) - (last(data.get(IORB, [])) or 0)) * 100, 1),
        round(((chg(data.get(SOFR, [])) or 0) - (chg(data.get(IORB, [])) or 0)) * 100, 1)
    )
    fills[39] = (
        round(((last(data.get(EFFR, [])) or 0) - (last(data.get(IORB, [])) or 0)) * 100, 1),
        round(((chg(data.get(EFFR, [])) or 0) - (chg(data.get(IORB, [])) or 0)) * 100, 1)
    )
    fills[40] = (
        round(((last(data.get(SOFR, [])) or 0) - (last(data.get(EFFR, [])) or 0)) * 100, 1),
        round(((chg(data.get(SOFR, [])) or 0) - (chg(data.get(EFFR, [])) or 0)) * 100, 1)
    )

    # A-Overnight
    fills[41] = (last(data.get(DFEDTARL, [])), None)
    fills[42] = (round(last(data.get(IORB, [])) or 0, 2), bp(chg(data.get(IORB, []))))
    fills[43] = (round(last(data.get(SOFR, [])) or 0, 2), bp(chg(data.get(SOFR, []))))

    # A-Liquidity
    rbal = data.get(WRESBAL, [])
    fills[47] = (round((last(rbal) or 0) / 1000, 2), round((chg(rbal) or 0) / 1000, 2))

    # B-OAS
    fills[50] = (bp(last(data.get(BAMLC0A0CM, []))), bp(chg(data.get(BAMLC0A0CM, []))))
    fills[51] = (bp(last(data.get(BAMLH0A0HYM2, []))), bp(chg(data.get(BAMLH0A0HYM2, []))))

    # B-Sovereign
    fills[58] = (bp(last(data.get(BAMLEMCBPIOAS, []))), bp(chg(data.get(BAMLEMCBPIOAS, []))))

    # B-Mortgage
    fills[66] = (round(last(data.get(MORTGAGE30US, [])) or 0, 2), bp(chg(data.get(MORTGAGE30US, []))))
    fills[67] = (bp(last(sp_mort)), bp(chg(sp_mort)))

    # --- Write cells ---
    count_today, count_delta = 0, 0
    for row_idx, (today_val, delta_val) in fills.items():
        if today_val is not None:
            ws.cell(row=row_idx, column=5, value=today_val)
            count_today += 1
        if delta_val is not None:
            ws.cell(row=row_idx, column=6, value=delta_val)
            count_delta += 1

    # Curve regime (R14) — text-only, no delta
    ws.cell(row=14, column=5, value=compute_curve_regime(data))

    # --- v3.5 Trigger condition status (R91 col G) ---
    hy = data.get(BAMLH0A0HYM2, [])
    hyg_data = data.get(HYG, [])
    fxy_data = data.get(FXY, [])
    spy_data = data.get(SPY, [])

    hy_d20 = bp(chg(hy, 20))
    hy_d5 = bp(chg(hy, 5))
    hyg_5d = ret_pct(hyg_data, 5)
    hyg_20d = ret_pct(hyg_data, 20)
    fxy_5d = ret_pct(fxy_data, 5)
    spy_ma = ma_n(spy_data, 200)
    spy_cur = last(spy_data)
    spy_below = spy_cur < spy_ma if (spy_cur and spy_ma) else None

    lines = []
    dd = "TRIGGERED" if (hy_d20 is not None and hy_d20 > 20) else "OK"
    lines.append(f"HY OAS 20dd={hy_d20}bp [{dd}]")
    lines.append(f"HY OAS 5dd={hy_d5}bp")
    if hyg_5d is not None:
        s = "TRIGGERED" if hyg_5d < -1.5 else "OK"
        lines.append(f"HYG 5d={hyg_5d}% [{s}]")
        lines.append(f"HYG DD%={ret_pct(hyg_data, len(hyg_data))}% max drawdown")
    if fxy_5d is not None:
        s = "TRIGGERED" if fxy_5d > 2.5 else "OK"
        lines.append(f"FXY 5d={fxy_5d}% [{s}]")
    if spy_below is not None:
        lines.append(f"SPY vs 200MA: {'BELOW!!' if spy_below else 'above'} ($SPY={spy_cur}, 200MA={spy_ma})")
    lines.append("Extreme: VIX>35•HYOAS>5%•SOFR-IORB>+5bp")
    ws.cell(row=91, column=6, value="\n".join(lines))

    return count_today, count_delta


def main():
    parser = argparse.ArgumentParser(description="Auto-fill v5.1 daily xlsx")
    parser.add_argument("--date", help="Fill date YYYY-MM-DD")
    args = parser.parse_args()

    fill_dt = args.date or date.today().isoformat()
    if not TEMPLATE_XLSX.exists():
        print(f"[ERROR] Template not found: {TEMPLATE_XLSX}", file=sys.stderr)
        sys.exit(1)

    data = load_data()
    wb = load_workbook(str(TEMPLATE_XLSX))
    ws = wb["Daily Input"]
    n_today, n_delta = fill_xlsx(ws, data, fill_dt)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1) Save snapshot to daily_logs/ for audit trail
    snapshot = OUTPUT_DIR / f"daily_{fill_dt}.xlsx"
    wb.save(str(snapshot))
    print(f"[SNAPSHOT] {snapshot}")

    # 2) Overwrite the working template in-place
    wb.save(str(TEMPLATE_XLSX))
    print(f"[TEMPLATE] {TEMPLATE_XLSX} (updated)")

    print(f"          Today={n_today} fields, 20ddelta={n_delta} fields")
    return 0


if __name__ == "__main__":
    sys.exit(main())
